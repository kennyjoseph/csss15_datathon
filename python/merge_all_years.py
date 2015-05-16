__author__ = 'kjoseph'

import glob
from openpyxl import load_workbook

all_events = open("../data/processed/all_2013_911_events.tsv","w")
all_events.write("Date\tType\tPriority\tDistrict\n")
for fil in glob.glob("../data/911/Evh_2013*"):
    input_file = load_workbook(filename = fil,
                                use_iterators = True)
    print 'loaded'
    input_sheet = input_file.worksheets[len(input_file.worksheets)-1]
    i = 0
    for row in input_sheet.iter_rows():
        val = [cell.value for cell in row]
        i+=1
        if i == 1 or None in val[1:]:
            continue

        all_events.write("\t".join([val[1].split(" ")[0]] + val[2:])+"\n")

all_events.close()