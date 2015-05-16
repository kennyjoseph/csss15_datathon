__author__ = 'kjoseph'

import codecs
from collections import defaultdict
from openpyxl import load_workbook


input_file = load_workbook(filename = "../data/911/CJP_CPD_ServiceCalls_2010_NoAdmin.xlsx",
                           use_iterators = True)
print 'loaded'
input_sheet = input_file.worksheets[len(input_file.worksheets)-1]

i= 0
output_fil = open("../data/processed/911_cat_to_full.tsv","w")
for row in input_sheet.iter_rows():
    val = [cell.value for cell in row]
    if val[0] is not None and 'Final Event Type' in val[0]:
        mapping = val[0].replace("Final Event Type (Both)-->","",1).strip()
        output_fil.write(mapping.replace(":","\t") + "\n")
output_fil.close()
